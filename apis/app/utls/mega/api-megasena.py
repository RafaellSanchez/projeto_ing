import requests
import pandas as pd
import os
import openpyxl
import numpy as np
import itertools
import re

from datetime import datetime

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

file = '/workspaces/projeto_ing/apis/setting/Mega-Sena(3).xlsx'
file_save = '/workspaces/projeto_ing/apis/app/tmp/'
wb = openpyxl.load_workbook(filename=file)

name = f'resultado_mega_{timestamp}.txt'
caminho = file_save + name
print(caminho)

dict_dfs = {}


for sheet in wb.sheetnames:
    sheet_data = []
    for row in wb[sheet].iter_rows(values_only=True):
        sheet_data.append(row)
    
    df = pd.DataFrame(sheet_data[1:], columns=sheet_data[0])
    print(df)
    df.to_csv(caminho, sep=';', index=False)
    print(f'arquivo salvo: {caminho}')
    
load = '/workspaces/projeto_ing/apis/app/tmp/resultado_mega_20241230_150630.txt'

df = pd.read_csv(load, sep=';', index_col=False)
df_conc = df['Concurso']

valores_filtrados = [valor for valor in df_conc.values.flatten() if 1 <= valor <= 60]
valores_array = np.array(valores_filtrados)
valores_mais_frequentes, contagens = np.unique(valores_array, return_counts=True)

indices_mais_frequentes = np.argsort(-contagens)[:24]
valores_mais_frequentes = valores_mais_frequentes[indices_mais_frequentes]

valores_mais_frequentes = np.floor(60*np.random.random((10,6))) 
print(valores_mais_frequentes)